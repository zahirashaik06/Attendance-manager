"""
Attendance Manager
------------------
Reads attendance from `attendance.xlsx`, calculates Present/Absent counts and
attendance %, flags students below a threshold, and writes a 'Summary' sheet.

Usage:
    pip install openpyxl
    python attendance_manager.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "attendance.xlsx"
OUTPUT_FILE = "attendance_report.xlsx"
THRESHOLD = 75.0  # % below this is flagged as low attendance


def main():
    wb = load_workbook(INPUT_FILE)
    ws = wb["Attendance"]

    # Read headers and find day columns (everything after "Student Name")
    headers = [c.value for c in ws[1]]
    day_cols = [i for i, h in enumerate(headers, start=1) if str(h).startswith("Day")]
    total_days = len(day_cols)

    # Build / reset Summary sheet
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    summary = wb.create_sheet("Summary")

    summary.append([
        "Roll No", "Student Name", "Present", "Absent",
        "Total Days", "Attendance %", "Status",
    ])

    low_count = 0
    rows_processed = 0

    for row in ws.iter_rows(min_row=2, values_only=False):
        roll = row[0].value
        name = row[1].value
        if roll is None or name is None:
            continue
        marks = [str(row[i - 1].value).strip().upper() for i in day_cols]
        present = marks.count("P")
        absent = marks.count("A")
        pct = (present / total_days * 100) if total_days else 0
        status = "Low Attendance" if pct < THRESHOLD else "Good"
        if status == "Low Attendance":
            low_count += 1
        summary.append([roll, name, present, absent, total_days, round(pct, 2), status])
        rows_processed += 1

    # ---- Style Summary ----
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    low_fill = PatternFill("solid", start_color="F8CBAD")
    good_fill = PatternFill("solid", start_color="C6EFCE")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for col in range(1, 8):
        c = summary.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    for r in range(2, rows_processed + 2):
        for col in range(1, 8):
            c = summary.cell(row=r, column=col)
            c.border = border
            c.font = Font(name="Arial")
            if col != 2:
                c.alignment = center
        # Percentage formatting
        summary.cell(row=r, column=6).number_format = "0.00"
        # Color status
        status_cell = summary.cell(row=r, column=7)
        status_cell.fill = low_fill if status_cell.value == "Low Attendance" else good_fill

    widths = [10, 22, 10, 10, 12, 14, 18]
    for i, w in enumerate(widths, start=1):
        summary.column_dimensions[get_column_letter(i)].width = w
    summary.freeze_panes = "A2"

    # ---- Overall stats block ----
    start = rows_processed + 4
    summary.cell(row=start, column=1, value="Overall Stats").font = Font(bold=True, size=12)
    stats = [
        ("Total Students", rows_processed),
        ("Total Days", total_days),
        (f"Students Below {THRESHOLD}%", low_count),
        ("Class Avg Attendance %",
         f"=ROUND(AVERAGE(F2:F{rows_processed + 1}),2)"),
    ]
    for i, (label, val) in enumerate(stats, start=start + 1):
        summary.cell(row=i, column=1, value=label).font = Font(bold=True)
        summary.cell(row=i, column=2, value=val)

    wb.save(OUTPUT_FILE)
    print(f"✅ Report saved to {OUTPUT_FILE}")
    print(f"   Students processed: {rows_processed}")
    print(f"   Low attendance (<{THRESHOLD}%): {low_count}")


if __name__ == "__main__":
    main()
